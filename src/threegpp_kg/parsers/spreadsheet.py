from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass
from pathlib import PurePosixPath
from typing import Any

from openpyxl import load_workbook

from ..config import ParserConfig
from ..constants import FORBIDDEN_WORKBOOK_SHEETS
from ..domain import TDoc
from ..ingestion.normalize import normalize_column, normalize_status, split_multi_value
from .documents import UnsafeDocumentError, UnsupportedDocumentError


@dataclass(frozen=True, slots=True)
class ParsedSheet:
    sheet_name: str
    header_row: int
    rows_seen: int
    tdocs: list[TDoc]


def parse_tdoc_workbook(content: bytes, meeting_id: str, source_url: str) -> list[ParsedSheet]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parsed: list[ParsedSheet] = []
    for sheet in workbook.worksheets:
        if sheet.title.strip().lower() in FORBIDDEN_WORKBOOK_SHEETS:
            continue
        rows = sheet.iter_rows(values_only=True)
        header_index, headers, buffered = _find_header(rows)
        if headers is None:
            continue
        tdocs: list[TDoc] = []
        rows_seen = 0
        for values in buffered:
            rows_seen += 1
            record = _record(headers, values)
            tdoc_id = _text(record.get("tdoc"))
            if not tdoc_id or not _looks_like_tdoc(tdoc_id):
                continue
            status_raw = _text(record.get("tdoc_status") or record.get("conclusion"))
            tdocs.append(
                TDoc(
                    id=tdoc_id,
                    meeting_id=meeting_id,
                    title=_text(record.get("title") or record.get("subject")),
                    source=_text(record.get("source")),
                    document_type=_text(record.get("type")),
                    purpose=_text(record.get("for")),
                    agenda_item=_text(record.get("agenda_item") or record.get("ai")),
                    agenda_description=_text(record.get("agenda_item_description")),
                    status=normalize_status(status_raw),
                    status_raw=status_raw,
                    abstract=_text(record.get("abstract")),
                    summary=_text(record.get("summary")),
                    discussion=_text(record.get("discussion") or record.get("secretary_remarks")),
                    conclusion_text=_text(record.get("conclusion")),
                    revised_from=_optional(record.get("is_revision_of")),
                    revised_to=_optional(record.get("revised_to")),
                    releases=split_multi_value(_text(record.get("release"))),
                    specifications=split_multi_value(_text(record.get("spec"))),
                    work_items=split_multi_value(_text(record.get("related_wis"))),
                    cr_number=_optional(record.get("cr")),
                    cr_revision=_optional(record.get("cr_revision")),
                    cr_category=_optional(record.get("cr_category")),
                    source_url=source_url,
                )
            )
        if tdocs:
            parsed.append(ParsedSheet(sheet.title, header_index, rows_seen, tdocs))
    workbook.close()
    return parsed


def parse_tdoc_workbook_package(
    content: bytes,
    filename: str,
    meeting_id: str,
    source_url: str,
    parser_config: ParserConfig | None = None,
) -> list[ParsedSheet]:
    """Parse a direct XLSX export or an XLSX wrapped in a 3GPP ZIP package."""
    suffix = PurePosixPath(filename.casefold()).suffix
    if suffix == ".xlsx":
        return parse_tdoc_workbook(content, meeting_id, source_url)
    if suffix != ".zip":
        raise UnsupportedDocumentError(f"unsupported TDoc index extension: {suffix or '<none>'}")

    limits = parser_config or ParserConfig()
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if not members:
                raise UnsafeDocumentError("empty TDoc index archive is quarantined")
            if len(members) > limits.max_archive_members:
                raise UnsafeDocumentError("TDoc index archive contains too many members")
            if sum(member.file_size for member in members) > limits.max_archive_uncompressed_bytes:
                raise UnsafeDocumentError("TDoc index archive expands beyond the safe limit")

            candidates: list[zipfile.ZipInfo] = []
            for member in members:
                if member.flag_bits & 0x1:
                    raise UnsafeDocumentError("encrypted TDoc index archives are rejected")
                normalized = member.filename.replace("\\", "/")
                parts = [part for part in normalized.split("/") if part]
                if normalized.startswith("/") or ".." in parts:
                    raise UnsafeDocumentError(
                        f"unsafe TDoc index archive member: {member.filename}"
                    )
                basename = PurePosixPath(normalized).name
                if (
                    not member.is_dir()
                    and not basename.startswith(("~$", "._"))
                    and PurePosixPath(normalized.casefold()).suffix == ".xlsx"
                ):
                    candidates.append(member)
            if not candidates:
                raise UnsupportedDocumentError("TDoc index archive contains no XLSX workbook")
            selected = max(candidates, key=_workbook_member_key)
            workbook = archive.read(selected)
    except zipfile.BadZipFile as exc:
        raise UnsafeDocumentError("invalid TDoc index archive") from exc
    return parse_tdoc_workbook(workbook, meeting_id, source_url)


def _workbook_member_key(member: zipfile.ZipInfo) -> tuple[int, str]:
    name = PurePosixPath(member.filename).name.casefold()
    priority = 2 if "index" in name else 1 if "tdoc" in name or "agenda" in name else 0
    return priority, name


def _find_header(rows: Any) -> tuple[int, list[str] | None, list[tuple[Any, ...]]]:
    buffered: list[tuple[Any, ...]] = []
    for row_number, values in enumerate(rows, start=1):
        normalized = [normalize_column(_text(value)) for value in values]
        if "tdoc" in normalized and ("title" in normalized or "subject" in normalized):
            buffered.extend(tuple(row) for row in rows)
            return row_number, normalized, buffered
        if row_number >= 250:
            break
    return 0, None, []


def _record(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: values[index] if index < len(values) else None
        for index, header in enumerate(headers)
        if header
    }


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _optional(value: Any) -> str | None:
    text = _text(value)
    return text if text and text != "-" else None


def _looks_like_tdoc(value: str) -> bool:
    compact = value.upper().replace(" ", "")
    return compact.startswith(("R2-", "R3-", "S2-", "C1-")) and any(
        char.isdigit() for char in compact
    )
