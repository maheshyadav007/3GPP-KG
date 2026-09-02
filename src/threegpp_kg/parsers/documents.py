from __future__ import annotations

import hashlib
import io
import os
import posixpath
import re
import shutil
import subprocess
import tempfile
import zipfile
from collections.abc import Iterable
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree

from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

from ..config import ParserConfig
from ..constants import FORBIDDEN_WORKBOOK_SHEETS, BlockKind
from ..domain import DocumentBlock


class UnsafeDocumentError(ValueError):
    pass


class UnsupportedDocumentError(ValueError):
    pass


class LegacyConversionError(UnsupportedDocumentError):
    pass


class NoExtractableTextError(ValueError):
    pass


_STRICT_OOXML_NAMESPACES = {
    b"http://purl.oclc.org/ooxml/drawingml/main": (
        b"http://schemas.openxmlformats.org/drawingml/2006/main"
    ),
    b"http://purl.oclc.org/ooxml/drawingml/picture": (
        b"http://schemas.openxmlformats.org/drawingml/2006/picture"
    ),
    b"http://purl.oclc.org/ooxml/drawingml/wordprocessingDrawing": (
        b"http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
    ),
    b"http://purl.oclc.org/ooxml/officeDocument/customProperties": (
        b"http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
    ),
    b"http://purl.oclc.org/ooxml/officeDocument/customXml": (
        b"http://schemas.openxmlformats.org/officeDocument/2006/customXml"
    ),
    b"http://purl.oclc.org/ooxml/officeDocument/docPropsVTypes": (
        b"http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"
    ),
    b"http://purl.oclc.org/ooxml/officeDocument/extendedProperties": (
        b"http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
    ),
    b"http://purl.oclc.org/ooxml/officeDocument/math": (
        b"http://schemas.openxmlformats.org/officeDocument/2006/math"
    ),
    b"http://purl.oclc.org/ooxml/officeDocument/relationships": (
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ),
    b"http://purl.oclc.org/ooxml/wordprocessingml/main": (
        b"http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    ),
}


def parse_document(
    content: bytes,
    filename: str,
    document_id: str,
    parser_config: ParserConfig | None = None,
) -> list[DocumentBlock]:
    suffix = PurePosixPath(filename.lower()).suffix
    if suffix == ".docx":
        return parse_docx(sanitize_office_package(content), document_id)
    if suffix == ".pdf":
        return parse_pdf(content, document_id)
    if suffix == ".pptx":
        return parse_pptx(sanitize_office_package(content), document_id)
    if suffix == ".xlsx":
        return parse_xlsx(
            sanitize_office_package(content),
            document_id,
            parser_config or ParserConfig(),
        )
    if suffix == ".zip":
        return parse_document_archive(content, document_id, parser_config)
    if suffix == ".docm":
        raise UnsafeDocumentError("macro-enabled Office document is quarantined")
    if suffix == ".doc":
        return parse_legacy_doc(content, document_id, parser_config)
    if suffix == ".txt":
        return parse_text(content, document_id)
    raise UnsupportedDocumentError(f"unsupported document extension: {suffix or '<none>'}")


def parse_document_archive(
    content: bytes,
    document_id: str,
    parser_config: ParserConfig | None = None,
) -> list[DocumentBlock]:
    return _parse_document_archive(content, document_id, parser_config, depth=0)


def _parse_document_archive(
    content: bytes,
    document_id: str,
    parser_config: ParserConfig | None,
    depth: int,
) -> list[DocumentBlock]:
    limits = parser_config or ParserConfig()
    if depth > limits.max_archive_depth:
        raise UnsafeDocumentError("nested document archive exceeds the safe depth")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if not members:
                raise UnsafeDocumentError("empty document archive is quarantined")
            if len(members) > limits.max_archive_members:
                raise UnsafeDocumentError("document archive contains too many members")
            if sum(member.file_size for member in members) > limits.max_archive_uncompressed_bytes:
                raise UnsafeDocumentError("document archive expands beyond the safe limit")
            parsed: list[DocumentBlock] = []
            legacy_seen = False
            supported_seen = False
            for member in members:
                if member.flag_bits & 0x1:
                    raise UnsafeDocumentError("encrypted document archives are rejected")
                normalized = member.filename.replace("\\", "/")
                parts = [part for part in normalized.split("/") if part]
                if normalized.startswith("/") or ".." in parts:
                    raise UnsafeDocumentError(f"unsafe document archive member: {member.filename}")
                if any(part == "__MACOSX" or part.startswith("._") for part in parts):
                    continue
                suffix = PurePosixPath(normalized.lower()).suffix
                basename = PurePosixPath(normalized).name.casefold()
                if suffix == ".xlsx" and basename.startswith(("tdoc_list_", "partlist_")):
                    # Meeting exports are ingested separately as authoritative metadata. Reports
                    # commonly bundle copies whose formatted used range spans an entire worksheet.
                    continue
                if suffix == ".doc":
                    legacy_seen = True
                    if parser_config is None or parser_config.legacy_converter == "disabled":
                        continue
                if (
                    suffix
                    not in {
                        ".doc",
                        ".docx",
                        ".docm",
                        ".pdf",
                        ".pptx",
                        ".txt",
                        ".xlsx",
                        ".zip",
                    }
                    or member.is_dir()
                ):
                    continue
                supported_seen = True
                payload = archive.read(member)
                if suffix == ".zip":
                    parsed.extend(
                        _parse_document_archive(
                            payload,
                            document_id,
                            parser_config,
                            depth=depth + 1,
                        )
                    )
                else:
                    parsed.extend(parse_document(payload, normalized, document_id, parser_config))
    except zipfile.BadZipFile as exc:
        raise UnsafeDocumentError("invalid document archive") from exc
    if not parsed and legacy_seen:
        raise UnsupportedDocumentError(
            "archive contains only legacy DOC files; isolated conversion is required"
        )
    if not parsed and supported_seen:
        raise NoExtractableTextError(
            "archive contains supported documents but no extractable text"
        )
    if not parsed:
        raise UnsupportedDocumentError("archive contains no supported document")
    return [
        _block(document_id, index, block.kind, block.text, block.section_path).model_copy(
            update={"table_row": block.table_row}
        )
        for index, block in enumerate(parsed)
    ]


def parse_legacy_doc(
    content: bytes,
    document_id: str,
    parser_config: ParserConfig | None,
) -> list[DocumentBlock]:
    if parser_config is None or parser_config.legacy_converter == "disabled":
        raise UnsupportedDocumentError(
            "legacy DOC parsing requires a configured isolated conversion worker"
        )
    with tempfile.TemporaryDirectory(prefix="threegpp-doc-") as directory:
        workspace = Path(directory)
        source = workspace / "input.doc"
        converted = workspace / "input.docx"
        source.write_bytes(content)
        command = _legacy_conversion_command(parser_config.legacy_converter, source, converted)
        environment = {
            "HOME": str(workspace),
            "PATH": os.environ.get("PATH", "/usr/bin:/bin"),
            "TMPDIR": str(workspace),
        }
        try:
            result = subprocess.run(
                command,
                cwd=workspace,
                env=environment,
                capture_output=True,
                check=False,
                timeout=parser_config.legacy_conversion_timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise LegacyConversionError("legacy DOC conversion timed out") from exc
        if result.returncode != 0 or not converted.is_file():
            raise LegacyConversionError("legacy DOC conversion failed")
        if converted.stat().st_size > parser_config.max_converted_bytes:
            raise UnsafeDocumentError("converted legacy DOC exceeds the configured limit")
        return parse_docx(sanitize_office_package(converted.read_bytes()), document_id)


def _legacy_conversion_command(
    converter: str,
    source: Path,
    output: Path,
) -> list[str]:
    if converter in {"auto", "libreoffice"}:
        executable = shutil.which("soffice") or shutil.which("libreoffice")
        if executable:
            return [
                executable,
                "--headless",
                "--convert-to",
                "docx",
                "--outdir",
                str(output.parent),
                str(source),
            ]
        if converter == "libreoffice":
            raise LegacyConversionError("LibreOffice converter is not installed")
    if converter in {"auto", "textutil"}:
        executable = shutil.which("textutil")
        if executable:
            return [
                executable,
                "-convert",
                "docx",
                "-output",
                str(output),
                "--",
                str(source),
            ]
        if converter == "textutil":
            raise LegacyConversionError("textutil converter is not installed")
    raise LegacyConversionError("no supported legacy DOC converter is installed")


def validate_office_package(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = set(archive.namelist())
            if any(name.lower().endswith("vbaproject.bin") for name in names):
                raise UnsafeDocumentError("macro-enabled Office packages are rejected")
            for name in names:
                if not name.lower().endswith(".rels"):
                    continue
                root = ElementTree.fromstring(archive.read(name))
                for relationship in root:
                    if relationship.attrib.get("TargetMode", "").lower() != "external":
                        continue
                    relationship_type = relationship.attrib.get("Type", "")
                    target = relationship.attrib.get("Target", "")
                    allowed = relationship_type.endswith(("/hyperlink", "/attachedTemplate"))
                    local_ole = relationship_type.endswith(
                        "/oleObject"
                    ) and target.lower().startswith("file:")
                    cid_image = relationship_type.endswith("/image") and target.lower().startswith(
                        "cid:"
                    )
                    if not allowed and not local_ole and not cid_image:
                        raise UnsafeDocumentError(
                            f"external relationship type is not allowed in {name}"
                        )
    except zipfile.BadZipFile as exc:
        raise UnsafeDocumentError("invalid Office package") from exc
    except ElementTree.ParseError as exc:
        raise UnsafeDocumentError("malformed Office relationship XML") from exc


def sanitize_office_package(content: bytes) -> bytes:
    """Remove inert links and normalize Strict OOXML in a disposable parser copy."""
    validate_office_package(content)
    output = io.BytesIO()
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as source, zipfile.ZipFile(output, "w") as target:
            package_names = set(source.namelist())
            for member in source.infolist():
                payload = source.read(member)
                if member.filename.lower().endswith((".xml", ".rels")):
                    payload = _normalize_strict_ooxml(payload)
                if member.filename.lower().endswith(".rels"):
                    root = ElementTree.fromstring(payload)
                    for relationship in list(root):
                        if relationship.attrib.get("TargetMode", "").lower() == "external":
                            root.remove(relationship)
                            continue
                        resolved = _relationship_target(
                            member.filename, relationship.attrib.get("Target", "")
                        )
                        if resolved and resolved not in package_names:
                            root.remove(relationship)
                    payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
                target.writestr(member, payload)
    except zipfile.BadZipFile as exc:
        raise UnsafeDocumentError("invalid Office package") from exc
    except ElementTree.ParseError as exc:
        raise UnsafeDocumentError("malformed Office relationship XML") from exc
    return output.getvalue()


def _normalize_strict_ooxml(payload: bytes) -> bytes:
    for strict_namespace, transitional_namespace in _STRICT_OOXML_NAMESPACES.items():
        payload = payload.replace(strict_namespace, transitional_namespace)
    return payload


def _relationship_target(relationship_file: str, target: str) -> str | None:
    if not target or target.startswith("#"):
        return None
    relationship_path = PurePosixPath(relationship_file)
    if relationship_path.parent.name == "_rels":
        source_directory = relationship_path.parent.parent
    else:
        source_directory = relationship_path.parent
    normalized = posixpath.normpath(
        target.lstrip("/")
        if target.startswith("/")
        else posixpath.join(str(source_directory), target)
    )
    return normalized.lstrip("./")


def parse_docx(content: bytes, document_id: str) -> list[DocumentBlock]:
    document = Document(io.BytesIO(content))
    blocks: list[DocumentBlock] = []
    section_path: list[str] = []
    for item in _iter_docx_blocks(document):
        if isinstance(item, Paragraph):
            text = item.text.strip()
            if not text:
                continue
            style = (item.style.name or "").lower() if item.style else ""
            kind = _paragraph_kind(text, style)
            if kind == BlockKind.HEADING:
                level = _heading_level(style)
                section_path = section_path[: max(0, level - 1)] + [text]
            blocks.append(_block(document_id, len(blocks), kind, text, section_path))
            continue
        for row_number, row in enumerate(item.rows):
            text = " | ".join(cell.text.strip() for cell in row.cells if cell.text.strip())
            if text:
                block = _block(document_id, len(blocks), BlockKind.TABLE_ROW, text, section_path)
                blocks.append(block.model_copy(update={"table_row": row_number}))
    return blocks


def _iter_docx_blocks(document: Any) -> Iterable[Paragraph | Table]:
    for child in document.element.body.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, document)
        elif child.tag == qn("w:tbl"):
            yield Table(child, document)


def parse_pdf(content: bytes, document_id: str) -> list[DocumentBlock]:
    reader = PdfReader(io.BytesIO(content))
    blocks: list[DocumentBlock] = []
    for page_number, page in enumerate(reader.pages, start=1):
        for text in _paragraphs(page.extract_text() or ""):
            section = [f"Page {page_number}"]
            blocks.append(_block(document_id, len(blocks), BlockKind.PARAGRAPH, text, section))
    return blocks


def parse_text(content: bytes, document_id: str) -> list[DocumentBlock]:
    if b"\x00" in content:
        raise UnsupportedDocumentError("binary content is not a plain-text document")
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1252")
    return [
        _block(document_id, index, BlockKind.PARAGRAPH, paragraph, [])
        for index, paragraph in enumerate(_paragraphs(text))
    ]


def parse_xlsx(
    content: bytes,
    document_id: str,
    parser_config: ParserConfig,
) -> list[DocumentBlock]:
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise UnsafeDocumentError("invalid Excel workbook") from exc

    blocks: list[DocumentBlock] = []
    rows_seen = 0
    try:
        for sheet in workbook.worksheets:
            normalized_title = sheet.title.strip().lower()
            if normalized_title in FORBIDDEN_WORKBOOK_SHEETS or sheet.sheet_state != "visible":
                continue
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                rows_seen += 1
                if rows_seen > parser_config.max_workbook_rows:
                    raise UnsafeDocumentError("Excel workbook exceeds the configured row limit")
                cells = [_spreadsheet_text(value) for value in values]
                text = " | ".join(cell for cell in cells if cell)
                if not text:
                    continue
                block = _block(
                    document_id,
                    len(blocks),
                    BlockKind.TABLE_ROW,
                    text,
                    [sheet.title],
                )
                blocks.append(block.model_copy(update={"table_row": row_number}))
                if len(blocks) > parser_config.max_document_blocks:
                    raise UnsafeDocumentError("Excel workbook exceeds the configured block limit")
    finally:
        workbook.close()
    return blocks


def _spreadsheet_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value).strip()


def parse_pptx(content: bytes, document_id: str) -> list[DocumentBlock]:
    blocks: list[DocumentBlock] = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            slide_names = sorted(
                (
                    name
                    for name in archive.namelist()
                    if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
                ),
                key=lambda name: int(re.search(r"\d+", PurePosixPath(name).name).group()),  # type: ignore[union-attr]
            )
            for slide_number, name in enumerate(slide_names, start=1):
                root = ElementTree.fromstring(archive.read(name))
                slide_label = f"Slide {slide_number}"
                title = ""
                for shape in root.findall(".//{*}sp"):
                    paragraphs = _presentation_paragraphs(shape)
                    if not paragraphs:
                        continue
                    placeholder = shape.find(".//{*}ph")
                    placeholder_type = (
                        placeholder.attrib.get("type", "") if placeholder is not None else ""
                    )
                    is_title = placeholder_type in {"title", "ctrTitle"} or not title
                    for paragraph_index, (text, is_list) in enumerate(paragraphs):
                        if is_title and paragraph_index == 0 and not title:
                            title = text
                            blocks.append(
                                _block(
                                    document_id,
                                    len(blocks),
                                    BlockKind.HEADING,
                                    text,
                                    [slide_label, text],
                                )
                            )
                            continue
                        section = [slide_label, title] if title else [slide_label]
                        kind = BlockKind.LIST_ITEM if is_list else _paragraph_kind(text, "")
                        blocks.append(_block(document_id, len(blocks), kind, text, section))
                section = [slide_label, title] if title else [slide_label]
                for row_number, row in enumerate(root.findall(".//{*}tbl/{*}tr")):
                    cells = [
                        " ".join(
                            text.text.strip()
                            for text in cell.findall(".//{*}t")
                            if text.text and text.text.strip()
                        )
                        for cell in row.findall("./{*}tc")
                    ]
                    row_text = " | ".join(cell for cell in cells if cell)
                    if row_text:
                        block = _block(
                            document_id,
                            len(blocks),
                            BlockKind.TABLE_ROW,
                            row_text,
                            section,
                        )
                        blocks.append(block.model_copy(update={"table_row": row_number}))
    except zipfile.BadZipFile as exc:
        raise UnsafeDocumentError("invalid PowerPoint package") from exc
    except ElementTree.ParseError as exc:
        raise UnsafeDocumentError("malformed PowerPoint slide XML") from exc
    return blocks


def _presentation_paragraphs(shape: ElementTree.Element) -> list[tuple[str, bool]]:
    paragraphs: list[tuple[str, bool]] = []
    for paragraph in shape.findall(".//{*}txBody/{*}p"):
        text = "".join(node.text or "" for node in paragraph.findall(".//{*}t")).strip()
        if not text:
            continue
        properties = paragraph.find("./{*}pPr")
        is_list = properties is not None and (
            "lvl" in properties.attrib
            or properties.find("./{*}buChar") is not None
            or properties.find("./{*}buAutoNum") is not None
        )
        paragraphs.append((text, is_list))
    return paragraphs


def _paragraphs(text: str) -> Iterable[str]:
    current: list[str] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            if current:
                yield " ".join(current)
                current = []
        else:
            current.append(line)
    if current:
        yield " ".join(current)


def _paragraph_kind(text: str, style: str) -> BlockKind:
    lowered = text.lower()
    if "heading" in style or style in {"title", "doc-title"}:
        return BlockKind.HEADING
    if style.startswith("list"):
        return BlockKind.LIST_ITEM
    if lowered.startswith("discussion:") or style == "comments":
        return BlockKind.DISCUSSION
    if lowered.startswith("agreement:") or style == "agreement":
        return BlockKind.AGREEMENT
    if lowered.startswith("conclusion:"):
        return BlockKind.CONCLUSION
    if lowered.startswith("note:"):
        return BlockKind.NOTE
    return BlockKind.PARAGRAPH


def _heading_level(style: str) -> int:
    for token in reversed(style.split()):
        if token.isdigit():
            return max(1, int(token))
    return 1


def _block(
    document_id: str,
    index: int,
    kind: BlockKind,
    text: str,
    section_path: list[str],
) -> DocumentBlock:
    digest = hashlib.sha256(f"{document_id}|{index}|{text}".encode()).hexdigest()[:20]
    return DocumentBlock(
        id=f"block-{digest}",
        document_id=document_id,
        index=index,
        kind=kind,
        text=text,
        section_path=list(section_path),
    )
